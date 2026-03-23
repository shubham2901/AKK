"""
Recipe Table Cleanup Script v3
==============================
Input:  recipes_v3_two_tables.xlsx (2 sheets: recipes, recipe_videos)
Output: recipes_v3_cleaned.xlsx (2 sheets: recipes, recipe_videos)

Changes:
1. Cuisine: Merge region into cuisine. Specific cuisines get parent tags (Andhra → Andhra | South Indian)
2. Drop: region column
3. Meal time: New column replacing meal_type. Values: Breakfast, Lunch, Dinner, Evening Tea, Brunch, Late Night
4. Recipe type: Multi-select from: Main Course, Bread, Rice, Dal, Accompaniment, Snack, Dessert, Beverage, Soup, Salad
5. Diet tags: Standardize to fixed enum
6. Breakfast Staple → Main Course, Rice Main → Rice | Main Course
"""

import pandas as pd
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# CONFIG — Update input/output paths as needed
# ============================================================
INPUT_FILE = "recipes_v3_two_tables.xlsx"
OUTPUT_FILE = "recipes_v3_cleaned.xlsx"

# ============================================================
# ENUMS — Master lists for each field
# ============================================================

VALID_MEAL_TIMES = {"Breakfast", "Lunch", "Dinner", "Evening Tea", "Brunch", "Late Night"}

VALID_RECIPE_TYPES = {
    "Main Course", "Bread", "Rice", "Dal", "Accompaniment",
    "Snack", "Dessert", "Beverage", "Soup", "Salad"
}

VALID_DIET_TAGS = {
    "Vegetarian", "Eggetarian", "Non-Vegetarian", "Vegan", "Jain", "Sattvic",
    "Gluten-Free", "Dairy-Free", "Nut-Free", "Sugar-Free",
    "High-Protein", "Low-Calorie", "Keto", "Diabetic-Friendly"
}

# Cuisine → parent region mapping
# Specific cuisines auto-inherit their parent broad region
CUISINE_PARENTS = {
    "Punjabi": ["North Indian"],
    "Rajasthani": ["North Indian"],
    "Awadhi": ["North Indian"],
    "Mughlai": ["North Indian"],
    "Kashmiri": ["North Indian"],
    "Bihari": ["North Indian", "East Indian"],
    "Sindhi": ["North Indian"],
    "Andhra": ["South Indian"],
    "Tamil": ["South Indian"],
    "Kerala": ["South Indian"],
    "Karnataka": ["South Indian"],
    "Udupi": ["South Indian", "Karnataka"],
    "Chettinad": ["South Indian", "Tamil"],
    "Mangalorean": ["South Indian", "Karnataka"],
    "Hyderabadi": ["South Indian"],
    "Bengali": ["East Indian"],
    "Gujarati": ["West Indian"],
    "Maharashtrian": ["West Indian"],
    "Konkani": ["West Indian"],
    "Goan": ["West Indian"],
    "Malvani": ["West Indian", "Maharashtrian"],
}

# All valid cuisine tokens (flat list)
VALID_CUISINES = {
    "North Indian", "South Indian", "East Indian", "West Indian",
    "Punjabi", "Rajasthani", "Awadhi", "Mughlai", "Kashmiri", "Bihari", "Sindhi",
    "Andhra", "Tamil", "Kerala", "Karnataka", "Udupi", "Chettinad", "Mangalorean",
    "Hyderabadi", "Bengali", "Gujarati", "Maharashtrian", "Konkani", "Goan",
    "Indo-Chinese", "Italian", "Mexican", "Chinese", "Korean", "Continental",
    "Fusion", "Other", "Pan-Indian", "Jain",
}

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def pipe_split(val):
    """Split pipe-separated string into cleaned list."""
    if pd.isna(val) or str(val).strip() == "":
        return []
    return [x.strip() for x in str(val).split("|") if x.strip()]


def pipe_join(lst):
    """Join list into pipe-separated string. Dedupes and preserves order."""
    seen = set()
    result = []
    for item in lst:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return " | ".join(result) if result else ""


def clean_cuisine(raw_cuisine):
    """
    Parse raw cuisine string, extract valid tokens, strip 'Street Food',
    add parent regions for specific cuisines.
    """
    tokens = pipe_split(raw_cuisine)

    # Remove non-cuisine labels
    drop_labels = {"Street Food", "Dessert", "Other", "Indian"}
    cleaned = []
    for t in tokens:
        # Fuzzy match to valid cuisines
        matched = None
        for vc in VALID_CUISINES:
            if t.lower() == vc.lower():
                matched = vc
                break
        if matched and matched not in drop_labels:
            cleaned.append(matched)

    # Add parent regions
    expanded = []
    for c in cleaned:
        expanded.append(c)
        if c in CUISINE_PARENTS:
            for parent in CUISINE_PARENTS[c]:
                expanded.append(parent)

    # If nothing survived, mark as Pan-Indian
    if not expanded:
        expanded = ["Pan-Indian"]

    return pipe_join(expanded)


def clean_diet_tags(raw_tags):
    """Standardize diet tags to valid enum values."""
    tokens = pipe_split(raw_tags)
    cleaned = []
    for t in tokens:
        t_lower = t.lower().strip()
        # Map known aliases
        if t_lower == "eggless":
            continue  # Redundant with Vegetarian
        if t_lower == "non-veg":
            cleaned.append("Non-Vegetarian")
            continue
        if t_lower in ("diabetic", "diabetic-friendly"):
            cleaned.append("Diabetic-Friendly")
            continue
        # Match to valid tags
        for valid in VALID_DIET_TAGS:
            if t_lower == valid.lower():
                cleaned.append(valid)
                break
    if not cleaned:
        cleaned = ["Vegetarian"]  # Default for Hebbar's Kitchen
    return pipe_join(cleaned)


def clean_meal_time(raw_meal_type):
    """
    Convert old meal_type to new meal_time using valid enum.
    Maps: Snacks/Appetizer → dropped (that's recipe_type territory),
    Evening Tea stays, Brunch stays.
    """
    tokens = pipe_split(raw_meal_type)
    cleaned = []
    for t in tokens:
        t_lower = t.lower().strip()
        if t_lower == "breakfast":
            cleaned.append("Breakfast")
        elif t_lower == "lunch":
            cleaned.append("Lunch")
        elif t_lower == "dinner":
            cleaned.append("Dinner")
        elif t_lower in ("evening tea", "tea time", "tea"):
            cleaned.append("Evening Tea")
        elif t_lower == "brunch":
            cleaned.append("Brunch")
        elif t_lower == "late night":
            cleaned.append("Late Night")
        # "Snacks", "Appetizer", "Side Dish", "Condiment", "Dessert", "Beverage"
        # are all recipe_type concerns, not meal_time — skip them

    # If nothing mapped, infer from recipe context
    if not cleaned:
        # Default: most Indian food is lunch/dinner appropriate
        cleaned = ["Lunch", "Dinner"]
    return pipe_join(cleaned)


def clean_recipe_type(raw_type, recipe_name):
    """
    Standardize recipe_type. Handle migrations:
    - Breakfast Staple → Main Course
    - Rice Main → Rice | Main Course
    - Dal / Lentil → Dal
    Also re-classify based on recipe name for known misclassifications.
    """
    name = str(recipe_name).lower()

    # Keyword-based overrides (fixes known misclassifications from v2)
    # Pani Puri was wrongly tagged as Bread (matched on 'puri')
    if re.search(r'\bpani\s*puri\b|\bgolgappa\b|\bpuchka\b', name):
        return "Snack"
    # Weight loss / protein powder type items
    if re.search(r'\b(protein powder|weight loss)\b', name):
        return "Beverage"

    raw = str(raw_type).strip()
    mapping = {
        "Breakfast Staple": "Main Course",
        "Rice Main": "Rice | Main Course",
        "Dal / Lentil": "Dal",
        "Snack": "Snack",
        "Main Course": "Main Course",
        "Bread": "Bread",
        "Rice": "Rice",
        "Accompaniment": "Accompaniment",
        "Dessert": "Dessert",
        "Beverage": "Beverage",
        "Soup": "Soup",
        "Salad": "Salad",
    }

    result = mapping.get(raw, raw)

    # Validate each token
    tokens = pipe_split(result)
    validated = [t for t in tokens if t in VALID_RECIPE_TYPES]
    if not validated:
        validated = ["Main Course"]
    return pipe_join(validated)


# ============================================================
# MAIN PROCESSING
# ============================================================

def process():
    # Load both sheets
    df_recipes = pd.read_excel(INPUT_FILE, sheet_name="recipes")
    df_videos = pd.read_excel(INPUT_FILE, sheet_name="recipe_videos")

    print(f"Loaded {len(df_recipes)} recipes, {len(df_videos)} videos")

    # --- Clean cuisine (merge region, add parents, drop junk) ---
    df_recipes["cuisine"] = df_recipes["cuisine"].apply(clean_cuisine)

    # --- Drop region column ---
    if "region" in df_recipes.columns:
        df_recipes = df_recipes.drop(columns=["region"])

    # --- Clean diet tags ---
    df_recipes["diet_tags"] = df_recipes["diet_tags"].apply(clean_diet_tags)

    # --- Clean recipe_type ---
    df_recipes["recipe_type"] = df_recipes.apply(
        lambda r: clean_recipe_type(r["recipe_type"], r["recipe_name_english"]), axis=1
    )

    # --- Create meal_time from old meal_type, then drop meal_type ---
    df_recipes["meal_time"] = df_recipes["meal_type"].apply(clean_meal_time)
    if "meal_type" in df_recipes.columns:
        df_recipes = df_recipes.drop(columns=["meal_type"])

    # --- Reorder columns ---
    desired_order = [
        "recipe_id",
        "recipe_name_english", "recipe_name_hindi", "recipe_name_tamil",
        "recipe_name_telugu", "recipe_name_marathi", "recipe_name_malayalam",
        "alternative_names_english",
        "diet_tags", "cuisine", "meal_time", "recipe_type",
        "is_verified", "popularity_score",
        "hero_image", "description",
        "prep_time_minutes", "cook_time_minutes", "difficulty",
        "spice_level", "serving_size", "seasonal_tags",
        "created_at", "updated_at", "user_rating",
    ]
    # Only keep columns that exist
    final_order = [c for c in desired_order if c in df_recipes.columns]
    df_recipes = df_recipes[final_order]

    # --- Print summary ---
    print(f"\n=== CUISINE distribution (top 15) ===")
    all_cuisines = df_recipes["cuisine"].apply(pipe_split).explode().value_counts()
    print(all_cuisines.head(15).to_string())

    print(f"\n=== DIET TAGS distribution (top 10) ===")
    all_diets = df_recipes["diet_tags"].apply(pipe_split).explode().value_counts()
    print(all_diets.head(10).to_string())

    print(f"\n=== MEAL TIME distribution ===")
    all_meals = df_recipes["meal_time"].apply(pipe_split).explode().value_counts()
    print(all_meals.to_string())

    print(f"\n=== RECIPE TYPE distribution ===")
    all_types = df_recipes["recipe_type"].apply(pipe_split).explode().value_counts()
    print(all_types.to_string())

    # ============================================================
    # WRITE TO EXCEL WITH FORMATTING
    # ============================================================
    wb = Workbook()
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3748")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def write_sheet(ws, df):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    cell.font = data_font
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    # Recipes sheet
    ws_r = wb.active
    ws_r.title = "recipes"
    write_sheet(ws_r, df_recipes)

    # Videos sheet (unchanged structure, just pass through)
    ws_v = wb.create_sheet("recipe_videos")
    write_sheet(ws_v, df_videos)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to {OUTPUT_FILE}")
    print(f"Recipes: {len(df_recipes)}, Videos: {len(df_videos)}")


if __name__ == "__main__":
    process()
