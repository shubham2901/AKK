#!/usr/bin/env python3
"""
Generate SQL INSERT statements from recipes_v4_llm_cleaned.xlsx.
Outputs JSON files with batched SQL statements for execution via Supabase MCP.
"""

import os
import json
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "recipes_v4_llm_cleaned.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "seed_batches")

def parse_pipe_list(val):
    if not val or not isinstance(val, str):
        return []
    return [s.strip() for s in val.split("|") if s.strip()]

def safe_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None

def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def safe_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).lower() in ("true", "1", "yes")

def sql_escape(val):
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"

def sql_array(items):
    if not items:
        return "'{}'"
    escaped = [s.replace("'", "''").replace('"', '\\"') for s in items]
    inner = ",".join(f'"{e}"' for e in escaped)
    return f"'{{{inner}}}'"

def sql_int(val):
    if val is None:
        return "NULL"
    return str(int(val))

def sql_float(val):
    if val is None:
        return "NULL"
    return str(float(val))

def sql_bool(val):
    return "TRUE" if val else "FALSE"

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    recipes = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = {headers[i]: row[i].value for i in range(len(headers))}
        
        recipe_id = safe_int(vals.get("recipe_id"))
        if recipe_id is None:
            continue

        name_en = (vals.get("recipe_name_english") or "").strip()
        if not name_en:
            continue

        recipes.append(vals)

    wb.close()
    print(f"Parsed {len(recipes)} recipes")

    BATCH_SIZE = 100
    batch_num = 0

    for i in range(0, len(recipes), BATCH_SIZE):
        batch = recipes[i:i + BATCH_SIZE]
        
        values_list = []
        for vals in batch:
            recipe_id = safe_int(vals.get("recipe_id"))
            values_list.append(
                f"({sql_int(recipe_id)}, "
                f"{sql_escape(vals.get('recipe_name_english'))}, "
                f"{sql_escape(vals.get('recipe_name_hindi'))}, "
                f"{sql_escape(vals.get('recipe_name_tamil'))}, "
                f"{sql_escape(vals.get('recipe_name_telugu'))}, "
                f"{sql_escape(vals.get('recipe_name_marathi'))}, "
                f"{sql_escape(vals.get('recipe_name_malayalam'))}, "
                f"{sql_escape(vals.get('alternative_names_english'))}, "
                f"{sql_array(parse_pipe_list(vals.get('diet_tags')))}, "
                f"{sql_array(parse_pipe_list(vals.get('cuisine')))}, "
                f"{sql_array(parse_pipe_list(vals.get('meal_time')))}, "
                f"{sql_escape(vals.get('recipe_type'))}, "
                f"{sql_bool(safe_bool(vals.get('is_verified')))}, "
                f"{sql_float(safe_float(vals.get('popularity_score')) or 0)}, "
                f"{sql_escape(vals.get('hero_image'))}, "
                f"{sql_escape(vals.get('image_path'))}, "
                f"{sql_escape(vals.get('description'))}, "
                f"{sql_int(safe_int(vals.get('prep_time_minutes')))}, "
                f"{sql_int(safe_int(vals.get('cook_time_minutes')))}, "
                f"{sql_escape(vals.get('difficulty'))}, "
                f"{sql_escape(vals.get('spice_level'))}, "
                f"{sql_int(safe_int(vals.get('serving_size')))}, "
                f"{sql_array(parse_pipe_list(vals.get('seasonal_tags')))}, "
                f"NULL)"
            )

        sql = (
            "INSERT INTO recipes (id, recipe_name_english, recipe_name_hindi, recipe_name_tamil, "
            "recipe_name_telugu, recipe_name_marathi, recipe_name_malayalam, alternative_names_english, "
            "diet_tags, cuisine, meal_time, recipe_type, is_verified, popularity_score, hero_image, "
            "image_path, description, prep_time_minutes, cook_time_minutes, difficulty, spice_level, "
            "serving_size, seasonal_tags, one_line_hook) VALUES\n"
            + ",\n".join(values_list) + ";"
        )

        out_path = os.path.join(OUTPUT_DIR, f"batch_{batch_num:03d}.sql")
        with open(out_path, "w") as f:
            f.write(sql)
        
        print(f"  Wrote batch {batch_num} ({len(batch)} recipes) -> {out_path}")
        batch_num += 1

    print(f"\nGenerated {batch_num} batch files in {OUTPUT_DIR}")
    print(f"Total recipes: {len(recipes)}")

if __name__ == "__main__":
    main()
