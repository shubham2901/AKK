#!/usr/bin/env python3
"""Generate SQL INSERT batches for recipe_videos from recipes_v3_two_tables.xlsx."""

import os
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "recipes_v3_two_tables.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "video_batches")

SOURCE_TO_CHANNEL = {
    "hebbar": "Hebbars Kitchen",
    "ranveer": "Ranveer Brar",
    "sanjeev": "Sanjeev Kapoor",
}

def sql_escape(val):
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"

def sql_int(val):
    if val is None:
        return "NULL"
    try:
        return str(int(val))
    except (ValueError, TypeError):
        return "NULL"

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["recipe_videos"]
    headers = [c.value for c in ws[1]]

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    videos = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = {headers[i]: row[i].value for i in range(len(headers))}
        rid = vals.get("recipe_id")
        vid = vals.get("video_id")
        if rid is None or vid is None:
            continue
        videos.append(vals)

    wb.close()
    print(f"Parsed {len(videos)} videos")

    BATCH_SIZE = 200
    batch_num = 0

    for i in range(0, len(videos), BATCH_SIZE):
        batch = videos[i : i + BATCH_SIZE]
        values_list = []
        for v in batch:
            source = v.get("source") or ""
            channel = SOURCE_TO_CHANNEL.get(source.lower(), source)
            values_list.append(
                f"({sql_int(v['recipe_id'])}, "
                f"{sql_escape(v['video_id'])}, "
                f"{sql_escape(channel)}, "
                f"{sql_escape(v.get('youtube_link'))}, "  # title = youtube link for now
                f"{sql_escape(v.get('thumbnail'))}, "
                f"{sql_escape(v.get('youtube_link'))}, "
                f"{sql_escape(v.get('web_recipe_link'))}, "
                f"{sql_int(v.get('view_count'))}, "
                f"{sql_int(v.get('like_count'))}, "
                f"NULL)"  # published_at
            )

        sql = (
            "INSERT INTO recipe_videos (recipe_id, video_id, channel_name, title, thumbnail, url, web_recipe_link, views, likes, published_at) VALUES\n"
            + ",\n".join(values_list)
            + "\nON CONFLICT (recipe_id, video_id) DO NOTHING;"
        )

        out_path = os.path.join(OUTPUT_DIR, f"vbatch_{batch_num:03d}.sql")
        with open(out_path, "w") as f:
            f.write(sql)

        print(f"  Wrote batch {batch_num} ({len(batch)} videos)")
        batch_num += 1

    print(f"\nGenerated {batch_num} batch files, total {len(videos)} videos")

if __name__ == "__main__":
    main()
