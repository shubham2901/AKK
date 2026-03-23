"""
Convert checkpoint JSON to final Excel.
Reads recipes_checkpoint.json + recipes_v3_two_tables.xlsx → recipes_v4_llm_cleaned.xlsx
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# CONFIG — update paths if needed
# ============================================================
CHECKPOINT_FILE = "recipes_checkpoint.json"
INPUT_FILE = "recipes_v3_two_tables.xlsx"
OUTPUT_FILE = "recipes_v4_llm_cleaned.xlsx"

# Parent cuisine inheritance
CUISINE_PARENTS = {
    "Punjabi": ["North Indian"], "Rajasthani": ["North Indian"],
    "Awadhi": ["North Indian"], "Mughlai": ["North Indian"],
    "Kashmiri": ["North Indian"], "Bihari": ["East Indian"],
    "Sindhi": ["North Indian"], "Andhra": ["South Indian"],
    "Tamil": ["South Indian"], "Kerala": ["South Indian"],
    "Karnataka": ["South Indian"], "Udupi": ["South Indian", "Karnataka"],
    "Chettinad": ["South Indian", "Tamil"], "Mangalorean": ["South Indian", "Karnataka"],
    "Hyderabadi": ["South Indian"], "Bengali": ["East Indian"],
    "Gujarati": ["West Indian"], "Maharashtrian": ["West Indian"],
    "Konkani": ["West Indian"], "Goan": ["West Indian"],
}

VALID_CUISINES = {
    "North Indian", "South Indian", "East Indian", "West Indian",
    "Punjabi", "Rajasthani", "Awadhi", "Mughlai", "Kashmiri", "Bihari", "Sindhi",
    "Andhra", "Tamil", "Kerala", "Karnataka", "Udupi", "Chettinad", "Mangalorean",
    "Hyderabadi", "Bengali", "Gujarati", "Maharashtrian", "Konkani", "Goan",
    "Indo-Chinese", "Italian", "Mexican", "Chinese", "Korean", "Continental",
    "Fusion", "Other", "Pan-Indian", "Jain",
}
VALID_DIETS = {
    "Vegetarian", "Eggetarian", "Non-Vegetarian", "Vegan", "Jain", "Sattvic",
    "Gluten-Free", "Dairy-Free", "Nut-Free", "Sugar-Free",
    "High-Protein", "Low-Calorie", "Keto", "Diabetic-Friendly",
}
VALID_MEALS = {"Breakfast", "Lunch", "Dinner", "Evening Tea", "Brunch", "Late Night"}
VALID_TYPES = {
    "Main Course", "Bread", "Rice", "Dal", "Accompaniment",
    "Snack", "Dessert", "Beverage", "Soup", "Salad",
}


def pipe_join(lst):
    seen = set()
    result = []
    for item in lst:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return " | ".join(result) if result else ""


def validate(values, valid_set):
    return [v for v in values if v in valid_set]


def add_parents(cuisines):
    expanded = list(cuisines)
    for c in cuisines:
        if c in CUISINE_PARENTS:
            expanded.extend(CUISINE_PARENTS[c])
    return expanded


def main():
    with open(CHECKPOINT_FILE, "r") as f:
        classified = json.load(f)
    print(f"Loaded {len(classified)} classifications from checkpoint")

    df_recipes = pd.read_excel(INPUT_FILE, sheet_name="recipes")
    df_videos = pd.read_excel(INPUT_FILE, sheet_name="recipe_videos")
    print(f"Loaded {len(df_recipes)} recipes, {len(df_videos)} videos")

    # Apply classifications
    for idx in range(len(df_recipes)):
        key = str(idx)
        if key in classified:
            c = classified[key]
            cuisine_vals = validate(c.get("cuisine", []), VALID_CUISINES)
            cuisine_vals = add_parents(cuisine_vals)
            diet_vals = validate(c.get("diet_tags", []), VALID_DIETS)
            meal_vals = validate(c.get("meal_time", []), VALID_MEALS)
            type_vals = validate(c.get("recipe_type", []), VALID_TYPES)

            df_recipes.at[idx, "cuisine"] = pipe_join(cuisine_vals) or "Pan-Indian"
            df_recipes.at[idx, "diet_tags"] = pipe_join(diet_vals) or "Vegetarian"
            df_recipes.at[idx, "meal_time"] = pipe_join(meal_vals) or "Lunch | Dinner"
            df_recipes.at[idx, "recipe_type"] = pipe_join(type_vals) or "Main Course"
        else:
            # Unclassified — use defaults
            df_recipes.at[idx, "meal_time"] = "Lunch | Dinner"
            if "recipe_type" not in df_recipes.columns or pd.isna(df_recipes.at[idx, "recipe_type"]):
                df_recipes.at[idx, "recipe_type"] = "Main Course"

    # Drop old columns, reorder
    if "region" in df_recipes.columns:
        df_recipes = df_recipes.drop(columns=["region"])
    if "meal_type" in df_recipes.columns:
        df_recipes = df_recipes.drop(columns=["meal_type"])

    desired_order = [
        "recipe_id", "recipe_name_english", "recipe_name_hindi", "recipe_name_tamil",
        "recipe_name_telugu", "recipe_name_marathi", "recipe_name_malayalam",
        "alternative_names_english", "diet_tags", "cuisine", "meal_time", "recipe_type",
        "is_verified", "popularity_score", "hero_image", "description",
        "prep_time_minutes", "cook_time_minutes", "difficulty",
        "spice_level", "serving_size", "seasonal_tags", "created_at", "updated_at", "user_rating",
    ]
    final_order = [c for c in desired_order if c in df_recipes.columns]
    df_recipes = df_recipes[final_order]

    # Print summary
    def explode_counts(col):
        return df_recipes[col].str.split(" \\| ").explode().str.strip().value_counts()

    print(f"\nCuisine (top 15):\n{explode_counts('cuisine').head(15).to_string()}")
    print(f"\nDiet Tags:\n{explode_counts('diet_tags').head(10).to_string()}")
    print(f"\nMeal Time:\n{explode_counts('meal_time').to_string()}")
    print(f"\nRecipe Type:\n{explode_counts('recipe_type').to_string()}")

    # Write Excel
    wb = Workbook()
    hfont = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="2D3748")
    dfont = Font(name="Arial", size=10)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    def write_sheet(ws, df):
        for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                if r == 1:
                    cell.font, cell.fill = hfont, hfill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    cell.font = dfont
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    ws_r = wb.active
    ws_r.title = "recipes"
    write_sheet(ws_r, df_recipes)

    ws_v = wb.create_sheet("recipe_videos")
    write_sheet(ws_v, df_videos)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to {OUTPUT_FILE}")
    print(f"Recipes: {len(df_recipes)}, Videos: {len(df_videos)}")
    print(f"Unclassified (used defaults): {len(df_recipes) - len(classified)}")


if __name__ == "__main__":
    main()
