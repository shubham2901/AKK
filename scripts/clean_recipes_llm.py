"""
Recipe Table Cleanup Script v4 — LLM-powered categorization
============================================================
Uses Gemini API to properly classify recipes into:
  - cuisine (multi-value, with parent regions)
  - diet_tags (multi-value, from fixed enum)
  - meal_time (multi-value, from fixed enum)
  - recipe_type (multi-value, from fixed enum)

Usage:
  export GEMINI_API_KEY="your-key-here"
  pip install pandas openpyxl requests
  python clean_recipes_llm.py

Input:  recipes_v3_two_tables.xlsx
Output: recipes_v4_llm_cleaned.xlsx
"""

import pandas as pd
import numpy as np
import requests
import json
import time
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# CONFIG
# ============================================================
INPUT_FILE = "recipes_v3_two_tables.xlsx"
OUTPUT_FILE = "recipes_v4_llm_cleaned.xlsx"
CHECKPOINT_FILE = "recipes_checkpoint.json"  # saves progress so you can resume on crash

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"

BATCH_SIZE = 25  # recipes per API call (~90 calls for 2249 recipes)
RETRY_LIMIT = 3
RETRY_DELAY = 5  # seconds between retries
RATE_LIMIT_DELAY = 1.5  # seconds between successful calls

# ============================================================
# ENUMS — included in every prompt so Gemini picks from these
# ============================================================
ENUMS = {
    "cuisine": [
        "North Indian", "South Indian", "East Indian", "West Indian",
        "Punjabi", "Rajasthani", "Awadhi", "Mughlai", "Kashmiri", "Bihari", "Sindhi",
        "Andhra", "Tamil", "Kerala", "Karnataka", "Udupi", "Chettinad", "Mangalorean",
        "Hyderabadi", "Bengali", "Gujarati", "Maharashtrian", "Konkani", "Goan",
        "Indo-Chinese", "Italian", "Mexican", "Chinese", "Korean", "Continental",
        "Fusion", "Pan-Indian",
    ],
    "diet_tags": [
        "Vegetarian", "Eggetarian", "Non-Vegetarian", "Vegan", "Jain", "Sattvic",
        "Gluten-Free", "Dairy-Free", "Nut-Free", "Sugar-Free",
        "High-Protein", "Low-Calorie", "Keto", "Diabetic-Friendly",
    ],
    "meal_time": [
        "Breakfast", "Lunch", "Dinner", "Evening Tea", "Brunch", "Late Night",
    ],
    "recipe_type": [
        "Main Course", "Bread", "Rice", "Dal", "Accompaniment",
        "Snack", "Dessert", "Beverage", "Soup", "Salad",
    ],
}

# Parent region inheritance — applied AFTER Gemini classification
CUISINE_PARENTS = {
    "Punjabi": ["North Indian"],
    "Rajasthani": ["North Indian"],
    "Awadhi": ["North Indian"],
    "Mughlai": ["North Indian"],
    "Kashmiri": ["North Indian"],
    "Bihari": ["East Indian"],
    "Sindhi": ["North Indian"],
    "Andhra": ["South Indian"],
    "Tamil": ["South Indian"],
    "Kerala": ["South Indian"],
    "Karnataka": ["South Indian"],
    "Udupi": ["South Indian", "Karnataka"],
    "Chettinad": ["South Indian", "Tamil"],
    "Mangalorean": ["South Indian", "Karnataka"],
    "Hyderabadi": ["South Indian"],
    "Bengali": ["East Indian"],
    "Gujarati": ["West Indian"],
    "Maharashtrian": ["West Indian"],
    "Konkani": ["West Indian"],
    "Goan": ["West Indian"],
}


# ============================================================
# PROMPT
# ============================================================
SYSTEM_PROMPT = """You are a culinary expert specializing in Indian food. You categorize Indian recipes precisely.

RULES:
1. Only use values from the provided enum lists. Never invent new values.
2. All fields are multi-value. Return arrays.
3. For cuisine: Pick the MOST SPECIFIC cuisine(s). Do NOT add broad regions like "North Indian" or "South Indian" — those get added automatically by the script. Only add them if the dish genuinely belongs to a broad region with no more specific origin (e.g., "Dal Makhani" is broadly North Indian, not specifically Punjabi).
4. For diet_tags: "Vegetarian" means no meat, fish, or eggs. If a recipe has eggs, use "Eggetarian" instead. Hebbar's Kitchen is a vegetarian channel — almost everything is Vegetarian. Only mark Vegan if it has no dairy at all. Add secondary tags (Gluten-Free, High-Protein, etc.) only when clearly applicable.
5. For meal_time: When can someone typically eat this in India? A samosa is Evening Tea. A biryani is Lunch/Dinner. Dosa is Breakfast/Dinner. Gulab Jamun has no specific meal time — use Lunch/Dinner as default for desserts.
6. For recipe_type: What IS this dish?
   - "Main Course" = a full dish you'd eat as the center of a meal (curries, sabzis, biryani, dosa, idli, etc.)
   - "Rice" = rice preparations. Biryani and Pulao are BOTH "Rice" AND "Main Course".
   - "Dal" = lentil/dal preparations including sambar and rasam
   - "Bread" = roti, naan, paratha, puri, etc.
   - "Snack" = fried snacks, chaats, finger food, pakoras, street food items
   - "Accompaniment" = chutneys, raita, pickles, papads — things you eat WITH a meal
   - "Dessert" = sweets, mithai, cakes, ice cream
   - "Beverage" = drinks, juices, shakes, lassi
   - "Soup" = soups, shorba
   - "Salad" = salads, koshimbir

Respond with ONLY a JSON array. No markdown, no backticks, no explanation. Each element must have:
  {"idx": <number>, "cuisine": [...], "diet_tags": [...], "meal_time": [...], "recipe_type": [...]}
"""


def build_user_prompt(batch):
    """Build the per-batch user prompt with recipe list."""
    lines = []
    lines.append("Categorize these recipes. Use ONLY values from these enums:\n")
    for field, values in ENUMS.items():
        lines.append(f"{field}: {json.dumps(values)}")
    lines.append("\nRecipes:")
    for idx, row in batch.iterrows():
        name = row["recipe_name_english"]
        alt = row.get("alternative_names_english", "")
        old_cuisine = row.get("cuisine", "")
        old_diet = row.get("diet_tags", "")
        old_meal = row.get("meal_type", "")
        old_type = row.get("recipe_type", "")
        lines.append(
            f'  {{"idx": {idx}, "name": "{name}", "alt_names": "{alt}", '
            f'"old_cuisine": "{old_cuisine}", "old_diet": "{old_diet}", '
            f'"old_meal_type": "{old_meal}", "old_recipe_type": "{old_type}"}}'
        )
    lines.append("\nReturn JSON array only. No markdown.")
    return "\n".join(lines)


# ============================================================
# GEMINI API CALL
# ============================================================
def call_gemini(user_prompt):
    """Call Gemini API and return parsed JSON array."""
    payload = {
        "contents": [
            {"role": "user", "parts": [{"text": SYSTEM_PROMPT + "\n\n" + user_prompt}]}
        ],
        "generationConfig": {
            "temperature": 0.1,  # low temp for consistent categorization
            "maxOutputTokens": 8192,
        },
    }

    for attempt in range(RETRY_LIMIT):
        try:
            resp = requests.post(GEMINI_URL, json=payload, timeout=60)

            if resp.status_code == 429:
                wait = RETRY_DELAY * (attempt + 2)
                print(f"    Rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue

            if resp.status_code != 200:
                print(f"    API error {resp.status_code}: {resp.text[:200]}")
                time.sleep(RETRY_DELAY)
                continue

            data = resp.json()
            text = data["candidates"][0]["content"]["parts"][0]["text"]

            # Strip markdown fences if Gemini adds them despite instructions
            text = text.strip()
            if text.startswith("```"):
                text = text.split("\n", 1)[1]  # remove first line
            if text.endswith("```"):
                text = text.rsplit("```", 1)[0]
            text = text.strip()

            return json.loads(text)

        except json.JSONDecodeError as e:
            print(f"    JSON parse error (attempt {attempt+1}): {e}")
            print(f"    Raw response: {text[:300]}")
            time.sleep(RETRY_DELAY)
        except Exception as e:
            print(f"    Error (attempt {attempt+1}): {e}")
            time.sleep(RETRY_DELAY)

    return None  # all retries failed


# ============================================================
# POST-PROCESSING
# ============================================================
def pipe_join(lst):
    """Dedupe list and join with pipe."""
    seen = set()
    result = []
    for item in lst:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return " | ".join(result) if result else ""


def add_cuisine_parents(cuisines):
    """Add parent regions for specific cuisines."""
    expanded = list(cuisines)
    for c in cuisines:
        if c in CUISINE_PARENTS:
            expanded.extend(CUISINE_PARENTS[c])
    return expanded


def validate_field(values, valid_set):
    """Keep only values that are in the valid enum set."""
    return [v for v in values if v in valid_set]


# ============================================================
# CHECKPOINT MANAGEMENT
# ============================================================
def load_checkpoint():
    """Load previously classified results so we can resume."""
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r") as f:
            data = json.load(f)
        print(f"Loaded checkpoint with {len(data)} classified recipes")
        return data
    return {}


def save_checkpoint(results):
    """Save classified results to disk."""
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(results, f)


# ============================================================
# MAIN
# ============================================================
def main():
    if not GEMINI_API_KEY:
        print("ERROR: Set GEMINI_API_KEY environment variable")
        print("  export GEMINI_API_KEY='your-key-here'")
        sys.exit(1)

    # Load data
    df_recipes = pd.read_excel(INPUT_FILE, sheet_name="recipes")
    df_videos = pd.read_excel(INPUT_FILE, sheet_name="recipe_videos")
    print(f"Loaded {len(df_recipes)} recipes, {len(df_videos)} videos")

    # Load any previous progress
    classified = load_checkpoint()  # dict of str(idx) -> {cuisine, diet_tags, meal_time, recipe_type}

    # Figure out which recipes still need classification
    all_indices = list(df_recipes.index)
    remaining = [i for i in all_indices if str(i) not in classified]
    print(f"Already classified: {len(classified)}, Remaining: {len(remaining)}")

    # Process in batches
    total_batches = (len(remaining) + BATCH_SIZE - 1) // BATCH_SIZE

    for batch_num in range(total_batches):
        start = batch_num * BATCH_SIZE
        end = min(start + BATCH_SIZE, len(remaining))
        batch_indices = remaining[start:end]
        batch = df_recipes.loc[batch_indices]

        print(f"\nBatch {batch_num+1}/{total_batches} ({len(batch_indices)} recipes)")

        user_prompt = build_user_prompt(batch)
        result = call_gemini(user_prompt)

        if result is None:
            print(f"  FAILED — skipping batch. Will use fallback rules for these.")
            continue

        # Store results
        for item in result:
            idx = str(item["idx"])
            classified[idx] = {
                "cuisine": item.get("cuisine", []),
                "diet_tags": item.get("diet_tags", []),
                "meal_time": item.get("meal_time", []),
                "recipe_type": item.get("recipe_type", []),
            }

        # Save checkpoint after every batch
        save_checkpoint(classified)
        print(f"  Classified {len(result)} recipes. Total: {len(classified)}/{len(df_recipes)}")

        # Rate limit
        time.sleep(RATE_LIMIT_DELAY)

    # ============================================================
    # APPLY CLASSIFICATIONS TO DATAFRAME
    # ============================================================
    print("\nApplying classifications...")

    cuisine_set = set(ENUMS["cuisine"])
    diet_set = set(ENUMS["diet_tags"])
    meal_set = set(ENUMS["meal_time"])
    type_set = set(ENUMS["recipe_type"])

    for idx in all_indices:
        key = str(idx)
        if key in classified:
            c = classified[key]
            # Validate and add parents
            cuisine_vals = validate_field(c["cuisine"], cuisine_set)
            cuisine_vals = add_cuisine_parents(cuisine_vals)
            diet_vals = validate_field(c["diet_tags"], diet_set)
            meal_vals = validate_field(c["meal_time"], meal_set)
            type_vals = validate_field(c["recipe_type"], type_set)

            df_recipes.at[idx, "cuisine"] = pipe_join(cuisine_vals)
            df_recipes.at[idx, "diet_tags"] = pipe_join(diet_vals) or "Vegetarian"
            df_recipes.at[idx, "meal_time"] = pipe_join(meal_vals) or "Lunch | Dinner"
            df_recipes.at[idx, "recipe_type"] = pipe_join(type_vals) or "Main Course"
        else:
            # Fallback: keep existing values, just clean formatting
            df_recipes.at[idx, "meal_time"] = df_recipes.at[idx].get("meal_type", "Lunch | Dinner")
            df_recipes.at[idx, "recipe_type"] = df_recipes.at[idx].get("recipe_type", "Main Course")

    # ============================================================
    # DROP OLD COLUMNS, REORDER
    # ============================================================
    if "region" in df_recipes.columns:
        df_recipes = df_recipes.drop(columns=["region"])
    if "meal_type" in df_recipes.columns:
        df_recipes = df_recipes.drop(columns=["meal_type"])

    desired_order = [
        "recipe_id",
        "recipe_name_english", "recipe_name_hindi", "recipe_name_tamil",
        "recipe_name_telugu", "recipe_name_marathi", "recipe_name_malayalam",
        "alternative_names_english",
        "diet_tags", "cuisine", "meal_time", "recipe_type",
        "is_verified", "popularity_score",
        "hero_image", "description",
        "prep_time_minutes", "cook_time_minutes", "difficulty",
        "spice_level", "serving_size", "seasonal_tags",
        "created_at", "updated_at", "user_rating",
    ]
    final_order = [c for c in desired_order if c in df_recipes.columns]
    df_recipes = df_recipes[final_order]

    # ============================================================
    # PRINT SUMMARY
    # ============================================================
    def explode_counts(col):
        return df_recipes[col].str.split(" \\| ").explode().str.strip().value_counts()

    print(f"\n{'='*50}")
    print("FINAL DISTRIBUTION SUMMARY")
    print(f"{'='*50}")
    print(f"\nCuisine (top 15):\n{explode_counts('cuisine').head(15).to_string()}")
    print(f"\nDiet Tags:\n{explode_counts('diet_tags').head(10).to_string()}")
    print(f"\nMeal Time:\n{explode_counts('meal_time').to_string()}")
    print(f"\nRecipe Type:\n{explode_counts('recipe_type').to_string()}")

    # ============================================================
    # WRITE EXCEL
    # ============================================================
    wb = Workbook()
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3748")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def write_sheet(ws, df):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    cell.font = data_font
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    ws_r = wb.active
    ws_r.title = "recipes"
    write_sheet(ws_r, df_recipes)

    ws_v = wb.create_sheet("recipe_videos")
    write_sheet(ws_v, df_videos)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to {OUTPUT_FILE}")
    print(f"Recipes: {len(df_recipes)}, Videos: {len(df_videos)}")

    # Clean up checkpoint on success
    print(f"\nCheckpoint file '{CHECKPOINT_FILE}' preserved — delete manually if you're happy with results.")


if __name__ == "__main__":
    main()
