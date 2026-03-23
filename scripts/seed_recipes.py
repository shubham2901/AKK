#!/usr/bin/env python3
"""
Seed the new recipes table from recipes_v4_llm_cleaned.xlsx.
Parses pipe-separated fields into Postgres arrays.
Also backfills one_line_hook from recipes_old where recipe names match.
"""

import os
import json
import openpyxl
from supabase import create_client

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://hrmhnovohubkfyxipjog.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

if not SUPABASE_KEY:
    raise ValueError("Set SUPABASE_SERVICE_KEY env var (service role key needed to bypass RLS for inserts)")

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "recipes_v4_llm_cleaned.xlsx")

def parse_pipe_list(val):
    """'Lunch | Dinner | Evening Tea' -> ['Lunch', 'Dinner', 'Evening Tea']"""
    if not val or not isinstance(val, str):
        return []
    return [s.strip() for s in val.split("|") if s.strip()]

def safe_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None

def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def safe_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).lower() in ("true", "1", "yes")

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Step 1: Fetch one_line_hook from old recipes table for backfill
    print("Fetching one_line_hook values from recipes_old...")
    hooks_resp = sb.from_("recipes_old").select("recipe_name_english, one_line_hook").not_.is_("one_line_hook", "null").execute()
    hooks_map = {}
    for row in hooks_resp.data:
        name = (row.get("recipe_name_english") or "").strip().lower()
        hook = row.get("one_line_hook")
        if name and hook:
            hooks_map[name] = hook
    print(f"  Found {len(hooks_map)} hooks to backfill")

    # Step 2: Parse Excel rows
    recipes = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = {headers[i]: row[i].value for i in range(len(headers))}

        recipe_id = safe_int(vals.get("recipe_id"))
        if recipe_id is None:
            continue

        name_en = (vals.get("recipe_name_english") or "").strip()
        if not name_en:
            continue

        hook_from_old = hooks_map.get(name_en.lower())

        recipes.append({
            "id": recipe_id,
            "recipe_name_english": name_en,
            "recipe_name_hindi": vals.get("recipe_name_hindi"),
            "recipe_name_tamil": vals.get("recipe_name_tamil"),
            "recipe_name_telugu": vals.get("recipe_name_telugu"),
            "recipe_name_marathi": vals.get("recipe_name_marathi"),
            "recipe_name_malayalam": vals.get("recipe_name_malayalam"),
            "alternative_names_english": vals.get("alternative_names_english"),
            "diet_tags": parse_pipe_list(vals.get("diet_tags")),
            "cuisine": parse_pipe_list(vals.get("cuisine")),
            "meal_time": parse_pipe_list(vals.get("meal_time")),
            "recipe_type": vals.get("recipe_type"),
            "is_verified": safe_bool(vals.get("is_verified")),
            "popularity_score": safe_float(vals.get("popularity_score")) or 0,
            "hero_image": vals.get("hero_image"),
            "image_path": vals.get("image_path"),
            "description": vals.get("description"),
            "prep_time_minutes": safe_int(vals.get("prep_time_minutes")),
            "cook_time_minutes": safe_int(vals.get("cook_time_minutes")),
            "difficulty": vals.get("difficulty"),
            "spice_level": vals.get("spice_level"),
            "serving_size": safe_int(vals.get("serving_size")),
            "seasonal_tags": parse_pipe_list(vals.get("seasonal_tags")),
            "one_line_hook": hook_from_old,
        })

    wb.close()
    print(f"Parsed {len(recipes)} recipes from Excel")

    # Step 3: Batch insert (Supabase client handles chunking)
    BATCH_SIZE = 200
    inserted = 0
    for i in range(0, len(recipes), BATCH_SIZE):
        batch = recipes[i : i + BATCH_SIZE]
        resp = sb.from_("recipes").insert(batch).execute()
        inserted += len(resp.data)
        print(f"  Inserted batch {i // BATCH_SIZE + 1}: {len(resp.data)} rows (total: {inserted})")

    print(f"\nDone! Inserted {inserted} recipes total.")

if __name__ == "__main__":
    main()
